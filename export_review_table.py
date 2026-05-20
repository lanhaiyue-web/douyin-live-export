"""Export Douyin live replay as one Claude-Code-style worksheet.

Columns:
    分钟 / 时间 / 在线人数 / 净进出 / 进入 / 离开 / 话术

Color:
    净进出 > 0: green
    净进出 < 0: red
    净进出 = 0: white
"""
from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from auth_browser import DEFAULT_PLATFORM, has_login_cookies, launch_persistent, resolve_platform
from douyin_sessions import (
    SESSION_CACHE_PATH,
    format_session_list,
    load_session_cache,
    refresh_sessions,
    select_cached_session,
)
from sources.anchor import (
    EP_MINUTE_TREND,
    REVIEW_URL,
    _dedupe_transcript,
    _is_transcript_url,
    _merge_minute_series,
    _minute_trend_url,
    _parse_minute,
    _parse_transcript_response,
    _transcript_fallback_time,
)
from sources.base import LiveSession, TranscriptSegment


HEADERS = ["分钟", "时间", "在线人数", "净进出", "进入", "离开", "话术"]


def desktop_path() -> Path:
    if os.name == "nt":
        try:
            import winreg

            key_path = r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path) as key:
                path, _ = winreg.QueryValueEx(key, "Desktop")
                return Path(os.path.expandvars(path))
        except Exception:
            pass
    return Path.home() / "Desktop"


def safe_filename(text: str, fallback: str = "直播复盘") -> str:
    text = (text or fallback).strip() or fallback
    text = re.sub(r'[\\/:*?"<>|\r\n]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:80] or fallback


def parse_room_id(url: str) -> str:
    qs = parse_qs(urlparse(url).query)
    return (qs.get("roomId") or qs.get("roomID") or qs.get("room_id") or [""])[0]


def parse_meta_from_text(text: str, room_id: str, url: str) -> dict[str, str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    def after(prefix: str) -> str:
        for idx, line in enumerate(lines):
            if line.startswith(prefix):
                rest = line.replace(prefix, "", 1).strip()
                return rest or (lines[idx + 1].strip() if idx + 1 < len(lines) else "")
        return ""

    title = ""
    try:
        idx = lines.index("经纪人：")
        title = lines[idx - 1]
    except Exception:
        title = ""

    return {
        "room_id": room_id,
        "page_url": url,
        "title": title or "直播复盘",
        "start_time": after("开播时间："),
        "end_time": after("关播时间："),
        "duration": after("直播时长："),
    }


def fill_meta_from_session(meta: dict[str, str], session: LiveSession | None) -> dict[str, str]:
    if not session:
        return meta
    meta = dict(meta)
    meta["room_id"] = meta.get("room_id") or session.session_id
    meta["title"] = (meta.get("title") if meta.get("title") != "直播复盘" else "") or session.title or "直播复盘"
    meta["start_time"] = meta.get("start_time") or session.start_time.strftime("%Y-%m-%d %H:%M:%S")
    meta["end_time"] = meta.get("end_time") or session.end_time.strftime("%Y-%m-%d %H:%M:%S")
    meta["duration"] = meta.get("duration") or f"{session.duration_min}分钟"
    return meta


def parse_transcript_records(text: str, start_time: datetime) -> list[TranscriptSegment]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    time_pat = re.compile(r"^20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$")
    records: list[TranscriptSegment] = []
    i = 0
    while i < len(lines):
        if not time_pat.match(lines[i]):
            i += 1
            continue

        ts_text = lines[i]
        speaker = lines[i - 1] if i > 0 else ""
        j = i + 1
        content: list[str] = []
        while j < len(lines):
            if j + 1 < len(lines) and time_pat.match(lines[j + 1]):
                break
            if time_pat.match(lines[j]):
                break
            content.append(lines[j])
            j += 1

        speech = " ".join(content).strip()
        if speech and (not speaker or "***" in speaker or "主播" in speaker):
            ts = datetime.strptime(ts_text, "%Y-%m-%d %H:%M:%S")
            minute = max(0, int((ts - start_time).total_seconds() // 60))
            records.append(TranscriptSegment(minute, ts, speech, source="anchor-dom"))
        i = max(j, i + 1)
    return records


def _transcript_list_info(page) -> dict[str, int] | None:
    return page.evaluate(
        """() => {
            const el = [...document.querySelectorAll('.ReactVirtualized__Grid.ReactVirtualized__List')]
              .find(e => (e.innerText || '').includes('20'));
            if (!el) return null;
            return {
                scrollTop: Math.round(el.scrollTop || 0),
                scrollHeight: Math.round(el.scrollHeight || 0),
                clientHeight: Math.round(el.clientHeight || 0),
                textLength: (el.innerText || '').length
            };
        }"""
    )


def _wait_transcript_list_ready(page, timeout_ms: int = 15_000) -> dict[str, int] | None:
    deadline = datetime.now().timestamp() + timeout_ms / 1000
    last_sig: tuple[int, int] | None = None
    stable_hits = 0
    info: dict[str, int] | None = None
    while datetime.now().timestamp() < deadline:
        info = _transcript_list_info(page)
        if info and info.get("scrollHeight", 0) > info.get("clientHeight", 0) > 0:
            sig = (int(info["scrollHeight"]), int(info["textLength"]))
            if sig == last_sig:
                stable_hits += 1
            else:
                stable_hits = 0
                last_sig = sig
            if stable_hits >= 2:
                return info
        page.wait_for_timeout(500)
    return info


def _scroll_transcript_list(page, pos: int) -> None:
    page.evaluate(
        """(pos) => {
            const el = [...document.querySelectorAll('.ReactVirtualized__Grid.ReactVirtualized__List')]
              .find(e => (e.innerText || '').includes('20'));
            if (!el) return;
            el.scrollTop = pos;
            el.dispatchEvent(new Event('scroll', {bubbles: true}));
        }""",
        pos,
    )


def _read_transcript_viewport(page, start_time: datetime) -> list[TranscriptSegment]:
    text = page.evaluate(
        """() => {
            const el = [...document.querySelectorAll('.ReactVirtualized__Grid.ReactVirtualized__List')]
              .find(e => (e.innerText || '').includes('20'));
            return el ? (el.innerText || '') : '';
        }"""
    )
    return parse_transcript_records(text, start_time)


def extract_transcript_from_current_page(page, start_time: datetime) -> list[TranscriptSegment]:
    """Scroll the virtualized transcript list on the current page."""
    info = _wait_transcript_list_ready(page)
    if not info:
        return []

    # ReactVirtualized renders asynchronously. A coarse, fast scroll can miss
    # the tail of the transcript, so use overlapping positions and wait for
    # each viewport to settle before reading innerText.
    seen: set[tuple[int, str, str]] = set()
    out: list[TranscriptSegment] = []

    def collect() -> None:
        for seg in _read_transcript_viewport(page, start_time):
            key = (seg.minute_index, seg.timestamp.isoformat(), seg.text)
            if key not in seen:
                seen.add(key)
                out.append(seg)

    step = max(160, min(360, int(info["clientHeight"] * 0.52)))
    for offset in (0, step // 2):
        pos = offset
        bottom_hits = 0
        while True:
            current = _transcript_list_info(page) or info
            max_scroll = max(0, int(current["scrollHeight"]) - int(current["clientHeight"]))
            pos = min(pos, max_scroll)
            _scroll_transcript_list(page, pos)
            page.wait_for_timeout(750)
            collect()
            current = _transcript_list_info(page) or current
            max_scroll = max(0, int(current["scrollHeight"]) - int(current["clientHeight"]))
            if pos >= max_scroll:
                bottom_hits += 1
                if bottom_hits >= 2:
                    break
                page.wait_for_timeout(1000)
            else:
                bottom_hits = 0
                pos = min(pos + step, max_scroll)

    # One last explicit bottom read catches late-rendered final records.
    info = _transcript_list_info(page) or info
    _scroll_transcript_list(page, int(info["scrollHeight"]))
    page.wait_for_timeout(1000)
    collect()

    out.sort(key=lambda seg: seg.timestamp)
    return out


def _has_large_internal_gap(segs: list[TranscriptSegment]) -> bool:
    mins = sorted({seg.minute_index for seg in segs if seg.text.strip()})
    if len(mins) < 3:
        return False
    return any((right - left) > 2 for left, right in zip(mins, mins[1:]))


def _tail_missing(segs: list[TranscriptSegment], expected_end_time: datetime | None) -> bool:
    if not expected_end_time or not segs:
        return False
    last_ts = max(seg.timestamp for seg in segs)
    return (expected_end_time - last_ts).total_seconds() > 120


def _video_duration(page) -> float:
    try:
        value = page.evaluate(
            """() => {
                const v = document.querySelector('video');
                return v && Number.isFinite(v.duration) ? v.duration : 0;
            }"""
        )
        return float(value or 0)
    except Exception:
        return 0.0


def _seek_video(page, seconds: float) -> None:
    try:
        page.evaluate(
            """(seconds) => {
                const v = document.querySelector('video');
                if (!v) return false;
                v.pause();
                v.currentTime = Math.max(0, Math.min(seconds, Number.isFinite(v.duration) ? v.duration : seconds));
                v.dispatchEvent(new Event('seeking'));
                v.dispatchEvent(new Event('timeupdate'));
                v.dispatchEvent(new Event('seeked'));
                return true;
            }""",
            seconds,
        )
    except Exception:
        return
    page.wait_for_timeout(4500)


def _video_checkpoints(duration: float, interval_seconds: int = 20 * 60) -> list[float]:
    if duration <= 0:
        return []
    points: set[int] = set()
    pos = interval_seconds
    while pos < duration - 60:
        points.add(pos)
        pos += interval_seconds
    points.add(max(0, int(duration - 180)))
    points.add(max(0, int(duration - 45)))
    return [float(p) for p in sorted(points) if p > 5]


def extract_transcript_with_video_sweep(
    page,
    start_time: datetime,
    expected_end_time: datetime | None = None,
) -> list[TranscriptSegment]:
    """Extract transcript by combining DOM scroll with video-time checkpoints.

    Douyin's transcript panel is tied to replay progress on some accounts. For
    longer lives, the first loaded panel can omit the tail or middle windows.
    We therefore collect once, then seek the replay video through coarse
    checkpoints and collect again until coverage looks complete.
    """
    collected = _dedupe_transcript(extract_transcript_from_current_page(page, start_time))
    if not _tail_missing(collected, expected_end_time) and not _has_large_internal_gap(collected):
        return collected

    duration = _video_duration(page)
    for checkpoint in _video_checkpoints(duration):
        _seek_video(page, checkpoint)
        collected = _dedupe_transcript(collected + extract_transcript_from_current_page(page, start_time))
        if not _tail_missing(collected, expected_end_time) and not _has_large_internal_gap(collected):
            break
    return collected


def extract_transcript_until_stable(
    page,
    start_time: datetime,
    expected_end_time: datetime | None = None,
) -> list[TranscriptSegment]:
    collected: list[TranscriptSegment] = []
    for attempt in range(3):
        collected = _dedupe_transcript(collected + extract_transcript_from_current_page(page, start_time))
        if not _has_large_internal_gap(collected) and not _tail_missing(collected, expected_end_time):
            break
        if attempt < 2:
            page.wait_for_timeout(2500)
    return collected


def _choose_page(context, room_id: str | None, current_page_only: bool = False):
    pages = list(context.pages)
    if room_id:
        for page in pages:
            if "anchor/review" in page.url and room_id in page.url:
                return page
    for page in pages:
        if "anchor/review" in page.url:
            return page
    if current_page_only:
        raise RuntimeError("托管浏览器里没有找到 anchor/review 复盘页。")
    if pages:
        return pages[0]
    return context.new_page()


def _click_text(page, text: str, timeout: int = 1500) -> bool:
    try:
        locator = page.get_by_text(text, exact=True).first
        locator.click(timeout=timeout)
        page.wait_for_timeout(600)
        return True
    except Exception:
        return False


def _trigger_review_panels(page) -> None:
    # These clicks trigger lazy XHR responses without opening or closing tabs.
    for label in ("营收", "流量", "互动指标", "粉丝"):
        _click_text(page, label, timeout=1200)
    for label in ("内容分析", "文字记录"):
        _click_text(page, label, timeout=2000)


def _try_direct_minute_fetch(page, session: LiveSession, traffic_by_minute: dict[str, dict[str, Any]]) -> None:
    try:
        body_text = page.evaluate(
            """async (url) => {
                const resp = await fetch(url, {credentials: 'include'});
                return await resp.text();
            }""",
            _minute_trend_url(session),
        )
        if body_text:
            _merge_minute_series(traffic_by_minute, json.loads(body_text))
    except Exception:
        pass


def capture_review_page(
    *,
    room_id: str | None = None,
    session: LiveSession | None = None,
    reload_page: bool = True,
    navigate: bool = True,
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    if session and not room_id:
        room_id = session.session_id

    traffic_by_minute: dict[str, dict[str, Any]] = {}
    transcript_bodies: list[tuple[str, dict[str, Any]]] = []

    spec = resolve_platform(DEFAULT_PLATFORM)
    with launch_persistent(spec) as context:
        if not has_login_cookies(context, spec):
            raise RuntimeError(
                f"持久 profile 未登录{spec.name_zh}。先跑：python auth_browser.py login {spec.key}"
            )
        page = _choose_page(context, room_id, current_page_only=not navigate)

        url_room_id = parse_room_id(page.url)
        if not room_id and url_room_id:
            room_id = url_room_id

        def on_response(resp) -> None:
            url = resp.url
            if EP_MINUTE_TREND in url:
                if room_id and room_id not in url:
                    return
                try:
                    _merge_minute_series(traffic_by_minute, resp.json())
                except Exception:
                    pass
            if _is_transcript_url(url):
                if room_id and room_id not in url:
                    return
                try:
                    transcript_bodies.append((url, resp.json()))
                except Exception:
                    pass

        page.on("response", on_response)

        if navigate and room_id:
            page.goto(f"{REVIEW_URL}?type=0&roomId={room_id}", wait_until="domcontentloaded", timeout=90_000)
            page.wait_for_timeout(5_000)
            # goto 之后再走一次 location.reload() 主环境触发，让抖音前端重发 minute_trend
            # （否则只跑 goto 时前端有时不会重发，监听不到流量数据）
            page.evaluate("() => { location.reload(); }")
            page.wait_for_timeout(15_000)
        elif reload_page:
            # page.reload() 走 CDP 协议，抖音前端能检测到自动化指纹返回空壳页（只剩导航 + 备案）。
            # 改成在主环境跑 location.reload()，抖音前端识别不出来，复盘内容能正常渲染。
            page.evaluate("() => { location.reload(); }")
            page.wait_for_timeout(15_000)
        else:
            page.wait_for_timeout(2_000)

        _trigger_review_panels(page)
        page.wait_for_timeout(3_000)

        body_text = page.locator("body").inner_text(timeout=10_000)
        room_id = room_id or parse_room_id(page.url)
        meta = parse_meta_from_text(body_text, room_id or "", page.url)
        meta = fill_meta_from_session(meta, session)
        if not meta["start_time"]:
            raise RuntimeError("当前页面没有读到开播时间，请确认托管浏览器已登录且复盘页加载完成。")

        start_time = datetime.strptime(meta["start_time"], "%Y-%m-%d %H:%M:%S")
        end_time = datetime.strptime(meta["end_time"], "%Y-%m-%d %H:%M:%S") if meta.get("end_time") else None
        if not session:
            session = LiveSession(
                source="anchor.douyin.com",
                session_id=meta["room_id"],
                title=meta["title"],
                start_time=start_time,
                end_time=end_time or start_time,
                duration_min=0,
            )
        _try_direct_minute_fetch(page, session, traffic_by_minute)

        transcript: list[TranscriptSegment] = []
        for url, body in transcript_bodies:
            transcript.extend(
                _parse_transcript_response(
                    body,
                    start_time,
                    fallback_time=_transcript_fallback_time(url, start_time),
                )
            )
        transcript.extend(extract_transcript_with_video_sweep(page, start_time, expected_end_time=end_time))
        transcript = _dedupe_transcript(transcript)

    if not traffic_by_minute:
        raise RuntimeError("没有捕获到 minute_trend。请确认选中的是目标复盘场次，并重新执行导出。")

    traffic = []
    prev_online = 0
    for idx, (_, raw) in enumerate(sorted(traffic_by_minute.items(), key=lambda item: item[0])):
        item = _parse_minute(raw, idx, prev_online)
        prev_online = item.online
        traffic.append(item)
    return meta, build_rows(traffic, transcript)


def build_rows(traffic, transcript: list[TranscriptSegment]) -> list[dict[str, Any]]:
    by_minute: dict[int, list[str]] = {}
    for seg in transcript:
        text = seg.text.strip()
        if not text:
            continue
        by_minute.setdefault(seg.minute_index, []).append(f"{seg.timestamp:%H:%M:%S} {text}")

    rows: list[dict[str, Any]] = []
    prev_online = None
    for idx, minute in enumerate(traffic):
        online = int(minute.online or 0)
        net = 0 if prev_online is None else online - prev_online
        prev_online = online
        rows.append(
            {
                "分钟": idx,
                "时间": minute.timestamp.strftime("%H:%M"),
                "在线人数": online,
                "净进出": net,
                "进入": int(minute.enter or 0),
                "离开": int(minute.leave or 0),
                "话术": "\n".join(by_minute.get(idx, [])),
            }
        )
    return rows


def load_rows_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "rows" in data:
        return data["rows"]
    if isinstance(data, list):
        return data
    raise ValueError(f"{path} 不是 rows JSON")


def export_rows(rows: list[dict[str, Any]], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "直播复盘"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="305496")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="E5E5E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
        excel_row = ws.max_row
        net = int(row.get("净进出") or 0)
        fill = green_fill if net > 0 else red_fill if net < 0 else white_fill
        for cell in ws[excel_row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 7))

    for idx, width in enumerate([10, 12, 14, 14, 12, 12, 120], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    for row_idx in range(2, ws.max_row + 1):
        text = ws.cell(row=row_idx, column=7).value or ""
        if text:
            ws.row_dimensions[row_idx].height = min(120, 18 + str(text).count("\n") * 18)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    target = out_path
    for idx in range(1, 100):
        try:
            wb.save(target)
            return target
        except PermissionError:
            target = out_path.with_name(f"{out_path.stem}_{idx + 1}{out_path.suffix}")
    raise PermissionError(f"无法保存 {out_path}，同名文件可能一直被占用。")


def default_out_path(meta: dict[str, str], title: str | None = None) -> Path:
    start_text = meta.get("start_time") or ""
    date_part = datetime.now().strftime("%Y-%m-%d")
    if start_text:
        try:
            date_part = datetime.strptime(start_text, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        except ValueError:
            pass
    name = f"{date_part}_{safe_filename(title or meta.get('title') or '直播复盘')}_流量话术复盘.xlsx"
    return desktop_path() / name


def ensure_session_cache(args) -> None:
    if args.refresh_sessions or not SESSION_CACHE_PATH.exists():
        sessions = refresh_sessions(limit=args.limit, auto_start=args.auto_start)
        print(f"CACHE={SESSION_CACHE_PATH}")
        print(format_session_list([json.loads(json.dumps(item, ensure_ascii=False)) for item in load_session_cache()]))
        if not sessions:
            raise RuntimeError("没有读取到直播场次。请确认托管浏览器已扫码登录。")


def command_list_sessions(args) -> None:
    if args.cache_only:
        items = load_session_cache()
    else:
        refresh_sessions(limit=args.limit, auto_start=args.auto_start)
        items = load_session_cache()
    print(f"CACHE={SESSION_CACHE_PATH}")
    print(format_session_list(items))


def main() -> None:
    parser = argparse.ArgumentParser(description="导出抖音直播复盘单表")
    parser.add_argument("--rows-json", type=Path, help="离线 rows JSON，直接生成 Excel")
    parser.add_argument("--current-page", action="store_true", help="从托管浏览器当前复盘页提取")
    parser.add_argument("--list-sessions", action="store_true", help="刷新/显示当前账号近 30 场缓存")
    parser.add_argument("--refresh-sessions", action="store_true", help="导出前强制刷新场次缓存")
    parser.add_argument("--cache-only", action="store_true", help="只读已有 sessions_cache.json，不碰浏览器")
    parser.add_argument("--limit", type=int, default=30, help="默认缓存近 30 场")
    parser.add_argument("--index", type=int, help="按 sessions_cache.json 中的序号导出")
    parser.add_argument("--room-id", help="按 room_id 导出")
    parser.add_argument("--title-contains", help="按标题关键词匹配缓存场次")
    parser.add_argument("--start-contains", help="按开播时间片段匹配缓存场次，如 2026-05-15")
    parser.add_argument("--reload", action="store_true", help="current-page 模式刷新当前页捕获数据")
    parser.add_argument("--no-reload", action="store_true", help="按场次导出时不刷新/重进页面")
    parser.add_argument("--auto-start", action="store_true", help="托管浏览器未运行时自动启动")
    parser.add_argument("--out", type=Path)
    args = parser.parse_args()

    meta: dict[str, str] = {"title": "直播复盘", "start_time": ""}

    if args.list_sessions:
        command_list_sessions(args)
        return

    if args.rows_json:
        rows = load_rows_json(args.rows_json)
        title = args.rows_json.stem.replace("_flow_transcript_rows", "")
    elif args.current_page:
        meta, rows = capture_review_page(room_id=args.room_id, reload_page=args.reload, navigate=False)
        title = meta.get("title") or "直播复盘"
    elif args.index or args.title_contains or args.start_contains:
        ensure_session_cache(args)
        session = select_cached_session(
            index=args.index,
            title_contains=args.title_contains,
            start_contains=args.start_contains,
        )
        meta, rows = capture_review_page(
            room_id=session.session_id,
            session=session,
            reload_page=not args.no_reload,
            navigate=True,
        )
        title = meta.get("title") or session.title
    elif args.room_id:
        session = None
        if SESSION_CACHE_PATH.exists():
            try:
                session = select_cached_session(room_id=args.room_id)
            except Exception:
                session = None
        meta, rows = capture_review_page(
            room_id=args.room_id,
            session=session,
            reload_page=not args.no_reload,
            navigate=True,
        )
        title = meta.get("title") or (session.title if session else "直播复盘")
    else:
        raise SystemExit(
            "必须指定一种模式：--list-sessions / --index N / --room-id ID / "
            "--title-contains 关键词 / --current-page / --rows-json 文件"
        )

    room_id = meta.get("room_id") or args.room_id or ""
    if room_id and not args.rows_json:
        raw_path = Path("data") / f"{room_id}_flow_transcript_rows.json"
        raw_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"ROWS_JSON={raw_path}")

    out = args.out or default_out_path(meta, title)
    saved = export_rows(rows, out)
    filled = sum(1 for row in rows if str(row.get("话术") or "").strip())
    print(f"EXCEL={saved}")
    print(f"ROWS={len(rows)}")
    print(f"TEXT_ROWS={filled}")


if __name__ == "__main__":
    main()
