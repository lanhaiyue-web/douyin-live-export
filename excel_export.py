"""把染色表格导出成 Excel 到用户桌面。

- 真实桌面位置从 Windows 注册表读取（用户改过位置也能正确找到）
- 文件名：{场次日期}直播复盘.xlsx，如 2026-05-15直播复盘.xlsx
- 染色保留：绿色 #C6EFCE，红色 #FFC7CE，白色不填
"""
import os
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analyzer import Row
from sources.base import LiveSession


GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

HEADERS = ["分钟", "时间", "在线人数", "净进出", "进入", "离开", "话术"]
COL_WIDTHS = {1: 8, 2: 10, 3: 10, 4: 10, 5: 10, 6: 10, 7: 70}


def get_desktop_path() -> Path:
    """读取真实桌面位置（兼容用户改过桌面路径的情况）。"""
    if os.name == "nt":
        try:
            import winreg
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path) as key:
                path, _ = winreg.QueryValueEx(key, "Desktop")
                return Path(os.path.expandvars(path))
        except Exception:
            pass
    return Path.home() / "Desktop"


def export_rows_to_excel(
    rows: List[Row],
    session: Optional[LiveSession] = None,
    summary_text: str = "",
    desktop_path: Optional[Path] = None,
) -> Path:
    """把染色表格 + Claude 总结导出成 Excel，返回保存路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "直播复盘"

    # 标题行
    ws.append(HEADERS)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行（按 color 染色）
    for r in rows:
        ws.append([
            r.minute,
            r.timestamp_str,
            r.online,
            r.net,
            r.enter,
            r.leave,
            r.text,
        ])
        excel_row = ws.max_row
        fill = None
        if r.color == "green":
            fill = GREEN_FILL
        elif r.color == "red":
            fill = RED_FILL
        if fill:
            for c in ws[excel_row]:
                c.fill = fill
        # 话术列自动换行
        ws.cell(row=excel_row, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    # 列宽 + 冻结首行
    for idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    # 第二个 sheet：Claude 总结
    if summary_text:
        ws2 = wb.create_sheet("Claude 复盘总结")
        ws2.column_dimensions["A"].width = 100
        for i, line in enumerate(summary_text.split("\n"), start=1):
            cell = ws2.cell(row=i, column=1, value=line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if line.startswith("#") or line.strip().endswith("：") or line.strip().endswith(":"):
                cell.font = Font(bold=True)

    # 文件名：用场次日期，没有就用今天
    if session and session.start_time:
        day = session.start_time.date()
    else:
        day = date.today()
    filename = f"{day.strftime('%Y-%m-%d')}直播复盘.xlsx"

    target_dir = desktop_path or get_desktop_path()
    target_dir.mkdir(parents=True, exist_ok=True)
    full_path = target_dir / filename

    # 如果同名文件已存在，追加 _2、_3...
    if full_path.exists():
        idx = 2
        while True:
            alt = target_dir / f"{day.strftime('%Y-%m-%d')}直播复盘_{idx}.xlsx"
            if not alt.exists():
                full_path = alt
                break
            idx += 1

    wb.save(full_path)
    return full_path
